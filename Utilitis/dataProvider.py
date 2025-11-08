import openpyxl


def get_data(sheetName, col_num):
    workbook = openpyxl.load_workbook("/Volumes/Entertainme/Automation/goibiboNewProject/Excel/testdata.xlsx")
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for row in range(2, totalrows+1):
        data = sheet.cell(row=row, column=col_num).value
        mainList.append([data])
    return mainList

