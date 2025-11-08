import openpyxl

workbook = openpyxl.load_workbook("/Volumes/Entertainme/Automation/goibiboNewProject/Excel/testdata.xlsx")
sheet = workbook["Login.test"]
totalrows = sheet.max_row
totalcols = sheet.max_column

print(f"total rows are :{totalrows}, total cols are: {totalcols}")
print(sheet.cell(row=2, column=1).value)

for rows in range(1, totalrows+1):
    for cols in range(1, totalcols+1):
        print(sheet.cell(row=rows, column=cols).value, end="   ")
    print()

